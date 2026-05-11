from datetime import datetime
from io import BytesIO
import re
from uuid import uuid4

from flask import Blueprint, request
from flask_jwt_extended import get_jwt_identity
from app.extensions import db
from app.models.import_batch import ImportBatch, StudentImportItem
from app.models.user import User
from app.models.reservation import Reservation, ReservationStatus
from app.models.room import Room
from app.utils import (
    admin_required,
    success_response,
    error_response,
    paginate_query,
    validate_email,
    validate_password,
    validate_student_id,
    validate_course_section,
)

admin_bp = Blueprint("admin", __name__, url_prefix="/api/admin")

def _gen_user_id(prefix: str = "AUTH") -> str:
    return f"{prefix}{uuid4().hex[:8].upper()}"

def _gen_id(prefix: str) -> str:
    return f"{prefix}{uuid4().hex[:10].upper()}"


@admin_bp.route("/metrics", methods=["GET"])
@admin_required
def metrics():
    """Lightweight admin dashboard summary."""
    users_total = User.query.filter(
        ~User.role.in_([User.db_role("admin"), User.db_role("superadmin")])
    ).count()

    reservations_total = Reservation.query.count()
    reservations_pending = Reservation.query.filter(
        Reservation.status == Reservation.db_status(ReservationStatus.PENDING)
    ).count()

    rooms_total = Room.query.filter(Room.is_active == 1).count()

    return success_response(
        {
            "users_total": users_total,
            "reservations_total": reservations_total,
            "reservations_pending": reservations_pending,
            "rooms_total": rooms_total,
        }
    )


def _norm_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _load_xlsx_workbook(file_storage):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return None, "Missing dependency: openpyxl. Install it to enable .xlsx uploads."

    try:
        raw = file_storage.read()
        wb = load_workbook(filename=BytesIO(raw), data_only=True)
        return wb, None
    except Exception as e:
        return None, f"Failed to read .xlsx file: {e}"

@admin_bp.route("/authorized-users", methods=["POST"])
@admin_required
def create_authorized_user():
    creator_id = get_jwt_identity()
    data = request.get_json(silent=True) or {}

    name           = (data.get("name")           or "").strip()
    email          = (data.get("email")          or "").strip().lower()
    password       =  data.get("password")       or ""
    course_section = (data.get("course_section") or "").strip().upper()

    errors = {}
    if not name:
        errors["name"] = "Name is required."
    if not email:
        errors["email"] = "Email is required."
    elif not validate_email(email):
        errors["email"] = "Must be a valid email address."
    if not password:
        errors["password"] = "Password is required."
    else:
        valid, msg = validate_password(password)
        if not valid:
            errors["password"] = msg
    if course_section and not validate_course_section(course_section):
        errors["course_section"] = (
            "Course/Section must follow the format AAAA X-X (e.g. BSIT 1-1)."
        )

    if errors:
        return error_response("Validation failed.", 422, errors)

    if User.query.filter_by(email=email).first():
        return error_response("Email is already registered.", 409)

    user = User(
        id=_gen_user_id("AUTH"),
        name=name,
        email=email,
        course_section=course_section or None,
        created_by=creator_id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    user.set_role("authorized_user")
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    return success_response(
        user.to_dict(include_sensitive=True),
        "Authorized user account created.", 201
    )


@admin_bp.route("/authorized-users", methods=["GET"])
@admin_required
def list_authorized_users():
    query = User.query.filter(User.role == User.db_role("authorized_user")).order_by(User.created_at.desc())

    is_active = request.args.get("is_active")
    if is_active is not None:
        query = query.filter_by(is_active=is_active.lower() == "true")

    search = request.args.get("search", "").strip()
    if search:
        query = query.filter(
            db.or_(User.name.ilike(f"%{search}%"), User.email.ilike(f"%{search}%"))
        )

    data = paginate_query(query, lambda u: u.to_dict(include_sensitive=True))
    return success_response(data)


@admin_bp.route("/authorized-users/<user_id>", methods=["GET"])
@admin_required
def get_authorized_user(user_id):
    user = User.query.filter(User.id == user_id, User.role == User.db_role("authorized_user")).first_or_404()
    return success_response(user.to_dict(include_sensitive=True))


@admin_bp.route("/authorized-users/<user_id>", methods=["PATCH"])
@admin_required
def update_authorized_user(user_id):
    user = User.query.filter(User.id == user_id, User.role == User.db_role("authorized_user")).first_or_404()
    data = request.get_json(silent=True) or {}

    if "name" in data and data["name"].strip():
        user.name = data["name"].strip()
    if "course_section" in data:
        new_cs = data["course_section"].strip().upper()
        if new_cs and not validate_course_section(new_cs):
            return error_response(
                "Course/Section must follow the format AAAA X-X (e.g. BSIT 1-1).", 422
            )
        user.course_section = new_cs or None
    if "is_active" in data:
        user.is_active = bool(data["is_active"])
    if "password" in data:
        valid, msg = validate_password(data["password"])
        if not valid:
            return error_response(msg, 422)
        user.set_password(data["password"])

    db.session.commit()
    return success_response(user.to_dict(include_sensitive=True), "Authorized user updated.")


@admin_bp.route("/authorized-users/<user_id>", methods=["DELETE"])
@admin_required
def deactivate_authorized_user(user_id):
    user = User.query.filter(User.id == user_id, User.role == User.db_role("authorized_user")).first_or_404()
    user.is_active = False
    db.session.commit()
    return success_response(message="Authorized user account deactivated.")

@admin_bp.route("/users", methods=["GET"])
@admin_required
def list_all_users():
    """View all users across all roles."""
    role  = request.args.get("role")
    query = User.query.filter(~User.role.in_([User.db_role("admin"), User.db_role("superadmin")])).order_by(User.created_at.desc())
    if role:
        role_key = (role or "").strip().lower()
        if role_key not in ("student", "authorized_user"):
            return error_response("Invalid role filter. Allowed: student, authorized_user.", 422)
        query = query.filter(User.role == User.db_role(role_key))

    search = request.args.get("search", "").strip()
    if search:
        query = query.filter(
            db.or_(User.name.ilike(f"%{search}%"), User.email.ilike(f"%{search}%"))
        )

    data = paginate_query(query, lambda u: u.to_dict(include_sensitive=True))
    return success_response(data)


@admin_bp.route("/users/<user_id>/activate", methods=["PATCH"])
@admin_required
def activate_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.role_key in ("admin", "superadmin"):
        return error_response("Admins and superadmins cannot be managed here.", 403)
    user.is_active = True
    db.session.commit()
    return success_response(message=f"User '{user.name}' activated.")


@admin_bp.route("/users/<user_id>/deactivate", methods=["PATCH"])
@admin_required
def deactivate_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.role_key in ("admin", "superadmin"):
        return error_response("Admins and superadmins cannot be managed here.", 403)
    user.is_active = False
    db.session.commit()
    return success_response(message=f"User '{user.name}' deactivated.")


@admin_bp.route("/users/<user_id>/role", methods=["PATCH"])
@admin_required
def update_user_role(user_id):
    """
    Promote/demote a user between 'student' and 'authorized_user'.

    Notes:
    - Admin/superadmin roles are managed by superadmin routes only.
    """
    data = request.get_json(silent=True) or {}
    role = (data.get("role") or "").strip()

    if role not in ("student", "authorized_user"):
        return error_response("Invalid role. Allowed: student, authorized_user.", 422)

    user = User.query.get_or_404(user_id)
    if user.role_key in ("admin", "superadmin"):
        return error_response("Admins and superadmins cannot be managed here.", 403)

    user.set_role(role)
    db.session.commit()
    return success_response(user.to_dict(include_sensitive=True), "User role updated.")


@admin_bp.route("/users", methods=["POST"])
@admin_required
def create_user():
    """Create a student/authorized_user account (admin-managed user CRUD)."""
    creator_id = get_jwt_identity()
    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    role = (data.get("role") or "student").strip().lower()
    student_id = (data.get("student_id") or "").strip()
    course_section = (data.get("course_section") or "").strip().upper()

    if role not in ("student", "authorized_user"):
        return error_response("Invalid role. Allowed: student, authorized_user.", 422)

    errors = {}
    if not name:
        errors["name"] = "Name is required."
    if not email:
        errors["email"] = "Email is required."
    elif not validate_email(email):
        errors["email"] = "Must be a valid email address."
    if not password:
        errors["password"] = "Password is required."
    else:
        valid, msg = validate_password(password)
        if not valid:
            errors["password"] = msg
    if student_id and not validate_student_id(student_id):
        errors["student_id"] = "Student ID must follow the format NN-NNNN (e.g. 21-0001)."
    if course_section and not validate_course_section(course_section):
        errors["course_section"] = "Course/Section must follow the format AAAA X-X (e.g. BSIT 1-1)."

    if errors:
        return error_response("Validation failed.", 422, errors)

    if User.query.filter_by(email=email).first():
        return error_response("Email is already registered.", 409)
    if student_id and User.query.filter_by(student_id=student_id).first():
        return error_response("Student ID is already registered.", 409)

    prefix = "STU" if role == "student" else "AUTH"
    user = User(
        id=_gen_user_id(prefix),
        name=name,
        email=email,
        student_id=student_id or None,
        created_by=creator_id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    if course_section:
        user.course_section = course_section
    user.set_role(role)
    user.set_password(password)

    db.session.add(user)
    db.session.commit()
    return success_response(user.to_dict(include_sensitive=True), "User account created.", 201)


@admin_bp.route("/users/<user_id>", methods=["GET"])
@admin_required
def get_user(user_id: str):
    user = User.query.get_or_404(user_id)
    return success_response(user.to_dict(include_sensitive=True))


@admin_bp.route("/users/<user_id>", methods=["PATCH"])
@admin_required
def update_user(user_id: str):
    """Update a student/authorized_user account."""
    updater_id = get_jwt_identity()
    user = User.query.get_or_404(user_id)

    if user.role_key in ("admin", "superadmin"):
        return error_response("Admins and superadmins cannot be managed here.", 403)

    data = request.get_json(silent=True) or {}

    if "name" in data and str(data["name"]).strip():
        user.name = str(data["name"]).strip()

    if "email" in data:
        new_email = (data.get("email") or "").strip().lower()
        if not new_email:
            return error_response("Email is required.", 422)
        if not validate_email(new_email):
            return error_response("Must be a valid email address.", 422)
        if User.query.filter(User.email == new_email, User.id != user.id).first():
            return error_response("Email is already registered.", 409)
        user.email = new_email

    if "student_id" in data:
        new_sid = (data.get("student_id") or "").strip()
        if new_sid and not validate_student_id(new_sid):
            return error_response("Student ID must follow the format NN-NNNN (e.g. 21-0001).", 422)
        if new_sid and User.query.filter(User.student_id == new_sid, User.id != user.id).first():
            return error_response("Student ID is already registered.", 409)
        user.student_id = new_sid or None

    if "course_section" in data:
        new_cs = (data.get("course_section") or "").strip().upper()
        if new_cs and not validate_course_section(new_cs):
            return error_response("Course/Section must follow the format AAAA X-X (e.g. BSIT 1-1).", 422)
        user.course_section = new_cs or None

    if "is_active" in data:
        user.is_active = bool(data.get("is_active"))

    if "role" in data:
        role = (data.get("role") or "").strip().lower()
        if role not in ("student", "authorized_user"):
            return error_response("Invalid role. Allowed: student, authorized_user.", 422)
        user.set_role(role)

    if "password" in data:
        valid, msg = validate_password(data.get("password") or "")
        if not valid:
            return error_response(msg, 422)
        user.set_password(data["password"])

    user.updated_by = updater_id
    user.updated_at = datetime.utcnow()
    db.session.commit()
    return success_response(user.to_dict(include_sensitive=True), "User updated.")


@admin_bp.route("/users/<user_id>", methods=["DELETE"])
@admin_required
def delete_user(user_id: str):
    """Soft-delete a student/authorized_user account (deactivate + deleted_at)."""
    user = User.query.get_or_404(user_id)
    if user.role_key in ("admin", "superadmin"):
        return error_response("Admins and superadmins cannot be managed here.", 403)

    user.is_active = False
    user.deleted_at = datetime.utcnow()
    user.updated_at = datetime.utcnow()
    db.session.commit()
    return success_response(message="User account deleted (deactivated).")


@admin_bp.route("/import/students", methods=["POST"])
@admin_required
def import_students_xlsx():
    # Admin .xlsx imports removed: keep this route blocked for backward compatibility.
    return error_response("Student import via .xlsx is available in Super Admin only.", 403)
    file = request.files.get("file")
    if not file:
      return error_response("file is required (.xlsx).", 422)
    if not str(file.filename or "").lower().endswith(".xlsx"):
        return error_response("Only .xlsx files are supported.", 422)

    default_password = (request.form.get("default_password") or "").strip()
    valid, msg = validate_password(default_password)
    if not valid:
        return error_response(f"default_password invalid: {msg}", 422)

    wb, err = _load_xlsx_workbook(file)
    if err:
        return error_response(err, 500)

    uploader_id = get_jwt_identity()
    batch = ImportBatch(
        id=_gen_id("IMP"),
        kind="students",
        filename=file.filename,
        created_by=uploader_id,
        created_at=datetime.utcnow(),
    )
    db.session.add(batch)

    created = 0
    updated = 0
    skipped = 0
    row_errors: list[dict] = []

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        col = {}
        for idx, row in enumerate(rows, start=1):
            if idx == 1:
                header = list(row or [])
                for j, h in enumerate(header):
                    key = _norm_header(h)
                    if key:
                        col[key] = j
                continue

            values = list(row or [])
            if not any(v is not None and str(v).strip() for v in values):
                continue

            def get(*keys):
                for k in keys:
                    j = col.get(k)
                    if j is None or j >= len(values):
                        continue
                    return values[j]
                return None

            name = str(get("name", "full_name", "student_name") or "").strip()
            email = str(get("email", "email_address") or "").strip().lower()
            student_id = str(get("student_id", "student_number", "student_no") or "").strip()
            course = str(get("course") or "").strip().upper()
            section = str(get("section") or "").strip().upper()
            course_section = str(get("course_section", "course_section_", "course_section__", "course_section___") or "").strip()
            role_raw = str(get("role", "account_role") or "").strip().lower()
            password = str(get("password") or "").strip()

            if not course_section and (course or section):
                course_section = f"{course} {section}".strip()

            role_key = role_raw if role_raw in ("student", "authorized_user") else "student"

            if not name or not email:
                skipped += 1
                row_errors.append({"sheet": ws.title, "row": idx, "error": "name and email are required."})
                continue

            if not validate_email(email):
                skipped += 1
                row_errors.append({"sheet": ws.title, "row": idx, "error": "invalid email."})
                continue

            if student_id and not validate_student_id(student_id):
                skipped += 1
                row_errors.append({"sheet": ws.title, "row": idx, "error": "invalid student_id format (NN-NNNN)."})
                continue

            if course_section and not validate_course_section(course_section.upper()):
                skipped += 1
                row_errors.append({"sheet": ws.title, "row": idx, "error": "invalid course_section format (AAAA X-X)."})
                continue

            existing = None
            if email:
                existing = User.query.filter_by(email=email).first()
            if not existing and student_id:
                existing = User.query.filter_by(student_id=student_id).first()

            if existing and existing.role_key in ("admin", "superadmin"):
                skipped += 1
                row_errors.append({"sheet": ws.title, "row": idx, "error": "cannot import over an admin/superadmin account."})
                continue

            chosen_pw = password or default_password
            valid_pw, msg_pw = validate_password(chosen_pw)
            if not valid_pw:
                skipped += 1
                row_errors.append({"sheet": ws.title, "row": idx, "error": f"invalid password: {msg_pw}"})
                continue

            if existing:
                if student_id and student_id != (existing.student_id or ""):
                    if User.query.filter(User.student_id == student_id, User.id != existing.id).first():
                        skipped += 1
                        row_errors.append({"sheet": ws.title, "row": idx, "error": "student_id already belongs to another user."})
                        continue

                existing.name = name
                existing.email = email
                existing.student_id = student_id or existing.student_id
                if course_section:
                    existing.course_section = course_section.upper()
                if role_key == "authorized_user" and existing.role_key == "student":
                    existing.set_role("authorized_user")

                # Update password only if explicitly provided in the sheet.
                if password:
                    existing.set_password(chosen_pw)

                existing.updated_by = uploader_id
                existing.updated_at = datetime.utcnow()

                db.session.add(StudentImportItem(batch_id=batch.id, user_id=existing.id, row_num=idx))
                updated += 1
                continue

            prefix = "AUTH" if role_key == "authorized_user" else "STU"
            u = User(
                id=_gen_user_id(prefix),
                name=name,
                email=email,
                student_id=student_id or None,
                created_by=uploader_id,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
                is_active=1,
            )
            if course_section:
                u.course_section = course_section.upper()
            u.set_role(role_key)
            u.set_password(chosen_pw)

            db.session.add(u)
            db.session.flush()
            db.session.add(StudentImportItem(batch_id=batch.id, user_id=u.id, row_num=idx))
            created += 1

    db.session.commit()

    return success_response(
        {
            "batch": batch.to_dict(),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": row_errors[:200],
        },
        "Student import complete.",
        201,
    )
