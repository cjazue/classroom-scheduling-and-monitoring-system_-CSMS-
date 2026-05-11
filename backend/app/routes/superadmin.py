from __future__ import annotations

from datetime import datetime, time
from io import BytesIO
import os
import re
from uuid import uuid4

from flask import Blueprint, request, send_file
from flask_jwt_extended import get_jwt_identity
from app.extensions import db
from app.models.reservation import Reservation, ReservationStatus
from app.models.schedule import Schedule
from app.models.import_batch import ImportBatch, ScheduleImportItem, StudentImportItem
from app.models.user import User
from app.utils import (
    superadmin_required,
    success_response,
    error_response,
    paginate_query,
    validate_email,
    validate_password,
    validate_student_id,
    validate_course_section,
)

superadmin_bp = Blueprint("superadmin", __name__, url_prefix="/api/superadmin")

def _gen_user_id(prefix: str = "AD") -> str:
    return f"{prefix}{uuid4().hex[:8].upper()}"


def _gen_id(prefix: str) -> str:
    return f"{prefix}{uuid4().hex[:10].upper()}"


_DAY_MAP = {
    "mon": "Monday",
    "monday": "Monday",
    "tue": "Tuesday",
    "tues": "Tuesday",
    "tuesday": "Tuesday",
    "wed": "Wednesday",
    "wednesday": "Wednesday",
    "thu": "Thursday",
    "thur": "Thursday",
    "thurs": "Thursday",
    "thursday": "Thursday",
    "fri": "Friday",
    "friday": "Friday",
    "sat": "Saturday",
    "saturday": "Saturday",
    "sun": "Sunday",
    "sunday": "Sunday",
}


def _normalize_day(value: object) -> str | None:
    if value is None:
        return None
    k = str(value).strip().lower()
    if not k:
        return None
    return _DAY_MAP.get(k, None)


_TIME_24H_RE = re.compile(r"^\s*(\d{1,2})\s*:\s*(\d{2})\s*$")
_TIME_12H_RE = re.compile(r"^\s*(\d{1,2})\s*:\s*(\d{2})\s*([AaPp][Mm])\s*$")


def _to_12h_str(hour: int, minute: int) -> str | None:
    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        return None
    ap = "AM" if hour < 12 else "PM"
    h = hour % 12
    if h == 0:
        h = 12
    return f"{h}:{minute:02d}{ap}"


def _coerce_schedule_time(value: object) -> str | None:
    """Accept Excel time, 24-hour strings, or 12-hour strings. Return 'H:MMAM/PM'."""
    if value is None:
        return None

    if isinstance(value, time):
        return _to_12h_str(value.hour, value.minute)

    s = str(value).strip()
    if not s:
        return None

    m12 = _TIME_12H_RE.match(s)
    if m12:
        hour = int(m12.group(1))
        minute = int(m12.group(2))
        ap = m12.group(3).lower()
        if hour < 1 or hour > 12 or minute < 0 or minute > 59:
            return None
        if ap == "am":
            hour = 0 if hour == 12 else hour
        else:
            hour = 12 if hour == 12 else hour + 12
        return _to_12h_str(hour, minute)

    m24 = _TIME_24H_RE.match(s)
    if m24:
        hour = int(m24.group(1))
        minute = int(m24.group(2))
        return _to_12h_str(hour, minute)

    return None


def _norm_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _load_xlsx_workbook(file_storage):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return None, "Missing dependency: openpyxl. Install it to enable .xlsx uploads."

    try:
        raw = file_storage.read()
        wb = load_workbook(filename=BytesIO(raw), data_only=True)
        return wb, None
    except Exception as e:
        return None, f"Failed to read .xlsx file: {e}"


def _xlsx_bytes_response(wb, download_name: str):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


def _make_import_template_workbook(sheet_title: str, headers: list[str], sample_rows: list[list[object]], notes: list[str]):
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return None, "Missing dependency: openpyxl. Install it to enable .xlsx templates."

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.append(headers)
    for row in sample_rows:
        ws.append(row)

    # Header style.
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="043AA3")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(45, len(str(header)) + 6))

    ws.freeze_panes = "A2"
    ws.auto_filter = f"A1:{get_column_letter(len(headers))}1"

    if notes:
        info = wb.create_sheet("Instructions")
        info.append(["CSMS Import Template Notes"])
        info["A1"].font = Font(bold=True)
        for n in notes:
            info.append([n])
        info.column_dimensions["A"].width = 140
        info.freeze_panes = "A2"

    return wb, None


@superadmin_bp.route("/import/templates/students.xlsx", methods=["GET"])
@superadmin_required
def download_student_import_template():
    headers = ["Name", "Email", "Student ID", "Course", "Section"]
    sample = [["Juan Dela Cruz", "juan.delacruz@example.com", "12-3456", "BSIT", "1-1"]]
    notes = [
        "Required columns: Name, Email.",
        "Optional columns: Student ID, Course, Section.",
        "Student ID format: NN-NNNN (example: 12-3456).",
        "Course + Section will be combined into Course Section (AAAA X-X) during import.",
        "Passwords are set automatically during import (no password column required).",
    ]

    wb, err = _make_import_template_workbook("Students", headers, sample, notes)
    if err:
        return error_response(err, 500)
    return _xlsx_bytes_response(wb, "csms_students_template.xlsx")


@superadmin_bp.route("/import/templates/schedules.xlsx", methods=["GET"])
@superadmin_required
def download_schedule_import_template():
    headers = ["Section", "Subject", "Subject Code", "Day", "Room", "Start Time", "End Time", "Campus", "Building"]
    sample = [["BSIT 1-1", "Programming 1", "IT101", "Monday", "RM101", "07:30", "09:00", "", ""]]
    notes = [
        "Required columns: Day, Room, Start Time, End Time.",
        "Section can be provided per row OR via the 'Section' form field in the upload UI.",
        "Section format: AAAA X-X (example: BSIT 1-1).",
        "Day accepts: Monday, Tue, Wed, Thu, Fri, Sat, Sun (case-insensitive).",
        "Time accepts: Excel time values, 24-hour HH:MM (07:30), or 12-hour H:MMAM/PM (7:30AM).",
        "Campus and Building are optional metadata fields.",
    ]

    wb, err = _make_import_template_workbook("Schedules", headers, sample, notes)
    if err:
        return error_response(err, 500)
    return _xlsx_bytes_response(wb, "csms_schedules_template.xlsx")

@superadmin_bp.route("/admins", methods=["POST"])
@superadmin_required
def create_admin():
    """Create an Admin account."""
    creator_id = get_jwt_identity()
    data = request.get_json(silent=True) or {}

    name           = (data.get("name")           or "").strip()
    email          = (data.get("email")          or "").strip().lower()
    password       =  data.get("password")       or ""
    course_section = (data.get("course_section") or "").strip()

    errors = {}
    if not name:
        errors["name"] = "Name is required."
    if not email:
        errors["email"] = "Email is required."
    elif not validate_email(email):
        errors["email"] = "Must be a valid email address."
    if not password:
        errors["password"] = "Password is required."
    else:
        valid, msg = validate_password(password)
        if not valid:
            errors["password"] = msg

    if errors:
        return error_response("Validation failed.", 422, errors)

    if User.query.filter_by(email=email).first():
        return error_response("Email is already registered.", 409)

    admin = User(
        id=_gen_user_id("AD"),
        name=name,
        email=email,
        course_section=course_section or None,
        created_by=creator_id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    admin.set_role("admin")
    admin.set_password(password)
    db.session.add(admin)
    db.session.commit()

    return success_response(admin.to_dict(include_sensitive=True), "Admin account created.", 201)


@superadmin_bp.route("/admins", methods=["GET"])
@superadmin_required
def list_admins():
    query = User.query.filter(User.role == User.db_role("admin")).order_by(User.created_at.desc())

    is_active = request.args.get("is_active")
    if is_active is not None:
        query = query.filter_by(is_active=is_active.lower() == "true")

    search = request.args.get("search", "").strip()
    if search:
        query = query.filter(
            db.or_(User.name.ilike(f"%{search}%"), User.email.ilike(f"%{search}%"))
        )

    data = paginate_query(query, lambda u: u.to_dict(include_sensitive=True))
    return success_response(data)


@superadmin_bp.route("/admins/<admin_id>", methods=["GET"])
@superadmin_required
def get_admin(admin_id):
    admin = User.query.filter(User.id == admin_id, User.role == User.db_role("admin")).first_or_404()
    return success_response(admin.to_dict(include_sensitive=True))


@superadmin_bp.route("/admins/<admin_id>", methods=["PATCH"])
@superadmin_required
def update_admin(admin_id):
    admin = User.query.filter(User.id == admin_id, User.role == User.db_role("admin")).first_or_404()
    data  = request.get_json(silent=True) or {}
    updater_id = get_jwt_identity()

    if "name" in data and data["name"].strip():
        admin.name = data["name"].strip()

    if "email" in data:
        new_email = (data.get("email") or "").strip().lower()
        if not new_email:
            return error_response("Email is required.", 422)
        if not validate_email(new_email):
            return error_response("Must be a valid email address.", 422)
        if User.query.filter(User.email == new_email, User.id != admin.id).first():
            return error_response("Email is already registered.", 409)
        admin.email = new_email

    if "course_section" in data:
        admin.course_section = data["course_section"].strip() or None
    if "is_active" in data:
        admin.is_active = bool(data["is_active"])
    if "password" in data:
        valid, msg = validate_password(data["password"])
        if not valid:
            return error_response(msg, 422)
        admin.set_password(data["password"])

    admin.updated_by = updater_id
    admin.updated_at = datetime.utcnow()
    db.session.commit()
    return success_response(admin.to_dict(include_sensitive=True), "Admin updated.")


@superadmin_bp.route("/admins/<admin_id>", methods=["DELETE"])
@superadmin_required
def delete_admin(admin_id):
    admin = User.query.filter(User.id == admin_id, User.role == User.db_role("admin")).first_or_404()
    admin.is_active = False
    admin.updated_at = datetime.utcnow()
    db.session.commit()
    return success_response(message="Admin account deactivated.")


@superadmin_bp.route("/metrics", methods=["GET"])
@superadmin_required
def metrics():
    """Lightweight system activity summary for dashboards."""
    users_total = User.query.count()

    by_role = {
        "superadmin": User.query.filter(User.role == User.db_role("superadmin")).count(),
        "admin": User.query.filter(User.role == User.db_role("admin")).count(),
        "authorized_user": User.query.filter(User.role == User.db_role("authorized_user")).count(),
        "student": User.query.filter(User.role == User.db_role("student")).count(),
    }

    reservations_total = Reservation.query.count()
    reservations_by_status = {
        "pending": Reservation.query.filter(
            Reservation.status == Reservation.db_status(ReservationStatus.PENDING)
        ).count(),
        "approved": Reservation.query.filter(
            Reservation.status == Reservation.db_status(ReservationStatus.APPROVED)
        ).count(),
        "rejected": Reservation.query.filter(
            Reservation.status == Reservation.db_status(ReservationStatus.REJECTED)
        ).count(),
        "cancelled": Reservation.query.filter(
            Reservation.status == Reservation.db_status(ReservationStatus.CANCELLED)
        ).count(),
    }

    return success_response(
        {
            "users_total": users_total,
            "users_by_role": by_role,
            "reservations_total": reservations_total,
            "reservations_by_status": reservations_by_status,
        }
    )


@superadmin_bp.route("/users", methods=["GET"])
@superadmin_required
def list_all_users():
    """Super Admin view of all users across all roles."""
    query = User.query.order_by(User.created_at.desc())

    role = (request.args.get("role") or "").strip().lower()
    if role:
        if role not in ("superadmin", "admin", "authorized_user", "student"):
            return error_response("Invalid role filter.", 422)
        query = query.filter(User.role == User.db_role(role))

    is_active = request.args.get("is_active")
    if is_active is not None:
        query = query.filter_by(is_active=is_active.lower() == "true")

    search = (request.args.get("search") or "").strip()
    if search:
        query = query.filter(
            db.or_(
                User.name.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%"),
                User.student_id.ilike(f"%{search}%"),
            )
        )

    data = paginate_query(query, lambda u: u.to_dict(include_sensitive=True))
    return success_response(data)


@superadmin_bp.route("/schedules", methods=["GET"])
@superadmin_required
def list_all_schedules():
    """Overall system view of schedules (all rows, not day-filtered)."""
    query = Schedule.query

    section = (request.args.get("section") or "").strip().upper()
    day = (request.args.get("day") or "").strip()
    room_key = (request.args.get("room_key") or "").strip()
    subject_code = (request.args.get("subject_code") or "").strip()
    subject = (request.args.get("subject") or "").strip()

    if section:
        query = query.filter(Schedule.section == section)
    if day:
        normalized = _normalize_day(day) or day
        query = query.filter(db.func.lower(Schedule.day) == str(normalized).lower())
    if room_key:
        query = query.filter(Schedule.room_key.ilike(f"%{room_key}%"))
    if subject_code:
        query = query.filter(Schedule.subject_code.ilike(f"%{subject_code}%"))
    if subject:
        query = query.filter(Schedule.subject.ilike(f"%{subject}%"))

    query = query.order_by(Schedule.section.asc(), Schedule.day.asc(), Schedule.start_time_raw.asc())
    data = paginate_query(query, lambda s: s.to_dict(include_room_key=True))
    return success_response(data)


@superadmin_bp.route("/schedules", methods=["POST"])
@superadmin_required
def create_schedule():
    data = request.get_json(silent=True) or {}

    section = (data.get("section") or "").strip().upper()
    day = _normalize_day(data.get("day"))
    room_key = (data.get("room_key") or "").strip()
    subject = (data.get("subject") or "").strip()
    subject_code = (data.get("subject_code") or "").strip()
    campus_id = (data.get("campus_id") or "").strip() or None
    building_id = (data.get("building_id") or "").strip() or None

    start_raw = _coerce_schedule_time(data.get("start_time"))
    end_raw = _coerce_schedule_time(data.get("end_time"))

    errors = {}
    if section and not validate_course_section(section):
        errors["section"] = "Section must follow the format AAAA X-X (e.g. BSIT 1-1)."
    if not section:
        errors["section"] = "Section is required."
    if not day:
        errors["day"] = "Day is required (e.g., Monday)."
    if not room_key:
        errors["room_key"] = "Room is required."
    if not start_raw:
        errors["start_time"] = "Invalid start_time. Use HH:MM (24h) or H:MMAM/PM."
    if not end_raw:
        errors["end_time"] = "Invalid end_time. Use HH:MM (24h) or H:MMAM/PM."

    if errors:
        return error_response("Validation failed.", 422, errors)

    # Basic time sanity check.
    tmp = Schedule(start_time_raw=start_raw, end_time_raw=end_raw)
    if tmp.start_minutes is None or tmp.end_minutes is None or tmp.start_minutes >= tmp.end_minutes:
        return error_response("start_time must be before end_time.", 422)

    s = Schedule(
        id=_gen_id("SCH"),
        section=section,
        subject=subject or None,
        subject_code=subject_code or None,
        day=day,
        campus_id=campus_id,
        building_id=building_id,
        room_key=room_key,
        start_time_raw=start_raw,
        end_time_raw=end_raw,
    )
    db.session.add(s)
    db.session.commit()
    return success_response(s.to_dict(include_room_key=True), "Schedule created.", 201)


@superadmin_bp.route("/schedules/<schedule_id>", methods=["GET"])
@superadmin_required
def get_schedule(schedule_id: str):
    s = Schedule.query.get_or_404(schedule_id)
    return success_response(s.to_dict(include_room_key=True))


@superadmin_bp.route("/schedules/<schedule_id>", methods=["PATCH"])
@superadmin_required
def update_schedule(schedule_id: str):
    s = Schedule.query.get_or_404(schedule_id)
    data = request.get_json(silent=True) or {}

    if "section" in data:
        section = (data.get("section") or "").strip().upper()
        if section and not validate_course_section(section):
            return error_response("Section must follow the format AAAA X-X (e.g. BSIT 1-1).", 422)
        s.section = section or None

    if "day" in data:
        day = _normalize_day(data.get("day"))
        if not day:
            return error_response("Invalid day. Use Monday..Sunday.", 422)
        s.day = day

    if "room_key" in data:
        rk = (data.get("room_key") or "").strip()
        if not rk:
            return error_response("room_key is required.", 422)
        s.room_key = rk

    if "subject" in data:
        s.subject = (data.get("subject") or "").strip() or None
    if "subject_code" in data:
        s.subject_code = (data.get("subject_code") or "").strip() or None
    if "campus_id" in data:
        s.campus_id = (data.get("campus_id") or "").strip() or None
    if "building_id" in data:
        s.building_id = (data.get("building_id") or "").strip() or None

    if "start_time" in data:
        start_raw = _coerce_schedule_time(data.get("start_time"))
        if not start_raw:
            return error_response("Invalid start_time. Use HH:MM (24h) or H:MMAM/PM.", 422)
        s.start_time_raw = start_raw

    if "end_time" in data:
        end_raw = _coerce_schedule_time(data.get("end_time"))
        if not end_raw:
            return error_response("Invalid end_time. Use HH:MM (24h) or H:MMAM/PM.", 422)
        s.end_time_raw = end_raw

    # Time sanity check when both present (or already set).
    if s.start_minutes is not None and s.end_minutes is not None and s.start_minutes >= s.end_minutes:
        return error_response("start_time must be before end_time.", 422)

    db.session.commit()
    return success_response(s.to_dict(include_room_key=True), "Schedule updated.")


@superadmin_bp.route("/schedules/<schedule_id>", methods=["DELETE"])
@superadmin_required
def delete_schedule(schedule_id: str):
    s = Schedule.query.get_or_404(schedule_id)
    # Best-effort cleanup of any import mapping.
    ScheduleImportItem.query.filter(ScheduleImportItem.schedule_id == s.id).delete()
    db.session.delete(s)
    db.session.commit()
    return success_response(message="Schedule deleted.")


@superadmin_bp.route("/imports", methods=["GET"])
@superadmin_required
def list_imports():
    kind = (request.args.get("kind") or "").strip().lower() or None
    query = ImportBatch.query.order_by(ImportBatch.created_at.desc())
    if kind:
        query = query.filter(ImportBatch.kind == kind)

    data = paginate_query(query, lambda b: b.to_dict())

    # Enrich counts (small N; OK to do per-row queries).
    for item in data["items"]:
        bid = item.get("id")
        if not bid:
            continue
        if item.get("kind") == "schedules":
            item["count"] = ScheduleImportItem.query.filter_by(batch_id=bid).count()
        elif item.get("kind") == "students":
            item["count"] = StudentImportItem.query.filter_by(batch_id=bid).count()

    return success_response(data)


@superadmin_bp.route("/imports/<batch_id>", methods=["GET"])
@superadmin_required
def get_import(batch_id: str):
    batch = ImportBatch.query.get_or_404(batch_id)
    out = batch.to_dict()
    if batch.kind == "schedules":
        out["count"] = ScheduleImportItem.query.filter_by(batch_id=batch_id).count()
    elif batch.kind == "students":
        out["count"] = StudentImportItem.query.filter_by(batch_id=batch_id).count()
    return success_response(out)


@superadmin_bp.route("/imports/<batch_id>", methods=["DELETE"])
@superadmin_required
def delete_import(batch_id: str):
    batch = ImportBatch.query.get_or_404(batch_id)

    if batch.kind == "schedules":
        schedule_ids = [r.schedule_id for r in ScheduleImportItem.query.filter_by(batch_id=batch_id).all()]
        if schedule_ids:
            Schedule.query.filter(Schedule.id.in_(schedule_ids)).delete(synchronize_session=False)
        ScheduleImportItem.query.filter_by(batch_id=batch_id).delete()
        db.session.delete(batch)
        db.session.commit()
        return success_response(message="Schedule import deleted.")

    if batch.kind == "students":
        user_ids = [r.user_id for r in StudentImportItem.query.filter_by(batch_id=batch_id).all()]
        if user_ids:
            # Soft-delete (deactivate) users affected by this import; keep history.
            User.query.filter(User.id.in_(user_ids)).update({"is_active": 0}, synchronize_session=False)
        StudentImportItem.query.filter_by(batch_id=batch_id).delete()
        db.session.delete(batch)
        db.session.commit()
        return success_response(message="Student import deleted (users deactivated).")

    return error_response("Unsupported import type.", 422)


@superadmin_bp.route("/import/students", methods=["POST"])
@superadmin_required
def import_students_xlsx():
    file = request.files.get("file")
    if not file:
        return error_response("file is required (.xlsx).", 422)
    if not str(file.filename or "").lower().endswith(".xlsx"):
        return error_response("Only .xlsx files are supported.", 422)

    default_password = (
        (request.form.get("default_password") or "").strip()
        or (os.environ.get("CSMS_DEFAULT_STUDENT_PASSWORD") or "").strip()
    )
    if not default_password:
        return error_response("default_password is required for imported students.", 422)
    valid, msg = validate_password(default_password)
    if not valid:
        return error_response(f"default_password invalid: {msg}", 422)

    wb, err = _load_xlsx_workbook(file)
    if err:
        return error_response(err, 500)

    uploader_id = get_jwt_identity()
    batch = ImportBatch(
        id=_gen_id("IMP"),
        kind="students",
        filename=file.filename,
        created_by=uploader_id,
        created_at=datetime.utcnow(),
    )
    db.session.add(batch)

    created = 0
    updated = 0
    skipped = 0
    row_errors: list[dict] = []

    def iter_sheets_rows():
        for ws in wb.worksheets:
            yield ws.title, ws.iter_rows(values_only=True)

    for sheet_name, rows in iter_sheets_rows():
        header = None
        col = {}
        for idx, row in enumerate(rows, start=1):
            if idx == 1:
                header = list(row or [])
                for j, h in enumerate(header):
                    key = _norm_header(h)
                    if key:
                        col[key] = j
                continue

            values = list(row or [])
            if not any(v is not None and str(v).strip() for v in values):
                continue

            def get(*keys):
                for k in keys:
                    j = col.get(k)
                    if j is None or j >= len(values):
                        continue
                    return values[j]
                return None

            first_name = str(get("first_name", "firstname", "first") or "").strip()
            second_name = str(get("second_name", "secondname", "second") or "").strip()
            middle_raw = str(get("middle_initial", "middle_name", "middlename", "middle", "mi") or "").strip()
            last_name = str(get("last_name", "lastname", "surname", "last") or "").strip()

            middle_initial = ""
            if middle_raw:
                # Always store/display Middle Initial (not full middle name).
                # Keep this conservative: take the first alphabetic character only.
                m = re.search(r"[A-Za-z]", middle_raw)
                middle_initial = m.group(0).upper() if m else ""

            composed_parts = [p for p in [first_name, second_name, (middle_initial or ""), last_name] if p]
            composed_name = " ".join(composed_parts).strip()

            name = composed_name or str(get("name", "full_name", "student_name") or "").strip()
            email = str(get("email", "email_address") or "").strip().lower()
            student_id = str(get("student_id", "student_number", "student_no") or "").strip()
            course = str(get("course") or "").strip().upper()
            section = str(get("section") or "").strip().upper()
            course_section = str(get("course_section", "course_section_", "course_section__", "course_section___") or "").strip()
            password = str(get("password") or "").strip()

            # If the sheet provides separate course/section, prefer them.
            if not course_section and (course or section):
                course_section = f"{course} {section}".strip()

            if not name or not email:
                skipped += 1
                row_errors.append(
                    {"sheet": sheet_name, "row": idx, "error": "name and email are required."}
                )
                continue

            if not validate_email(email):
                skipped += 1
                row_errors.append({"sheet": sheet_name, "row": idx, "error": "invalid email."})
                continue

            if student_id and not validate_student_id(student_id):
                skipped += 1
                row_errors.append({"sheet": sheet_name, "row": idx, "error": "invalid student_id format (NN-NNNN)."})
                continue

            if course_section and not validate_course_section(course_section.upper()):
                skipped += 1
                row_errors.append({"sheet": sheet_name, "row": idx, "error": "invalid course_section format (AAAA X-X)."})
                continue

            existing = None
            if email:
                existing = User.query.filter_by(email=email).first()
            if not existing and student_id:
                existing = User.query.filter_by(student_id=student_id).first()

            if existing and existing.role_key in ("admin", "superadmin"):
                skipped += 1
                row_errors.append({"sheet": sheet_name, "row": idx, "error": "cannot import over an admin/superadmin account."})
                continue

            if existing:
                # Avoid student_id collisions on update.
                if student_id and student_id != (existing.student_id or ""):
                    if User.query.filter(User.student_id == student_id, User.id != existing.id).first():
                        skipped += 1
                        row_errors.append({"sheet": sheet_name, "row": idx, "error": "student_id already belongs to another user."})
                        continue

                existing.name = name
                existing.email = email
                existing.student_id = student_id or existing.student_id
                if course_section:
                    existing.course_section = course_section.upper()
                # Do not demote previously upgraded users (authorized_user) on re-import.
                if existing.role_key == "student":
                    existing.set_role("student")
                existing.updated_by = uploader_id
                existing.updated_at = datetime.utcnow()

                db.session.add(StudentImportItem(batch_id=batch.id, user_id=existing.id, row_num=idx))
                updated += 1
                continue

            u = User(
                id=_gen_user_id("STU"),
                name=name,
                email=email,
                student_id=student_id or None,
                created_by=uploader_id,
                created_at=datetime.utcnow(),
                updated_at=datetime.utcnow(),
                is_active=1,
            )
            if course_section:
                u.course_section = course_section.upper()
            u.set_role("student")

            chosen_pw = password or default_password
            valid_pw, msg_pw = validate_password(chosen_pw)
            if not valid_pw:
                skipped += 1
                row_errors.append({"sheet": sheet_name, "row": idx, "error": f"invalid password: {msg_pw}"})
                continue
            u.set_password(chosen_pw)

            db.session.add(u)
            db.session.flush()
            db.session.add(StudentImportItem(batch_id=batch.id, user_id=u.id, row_num=idx))
            created += 1

    db.session.commit()

    return success_response(
        {
            "batch": batch.to_dict(),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": row_errors[:200],
        },
        "Student import complete.",
        201,
    )


@superadmin_bp.route("/import/schedules", methods=["POST"])
@superadmin_required
def import_schedules_xlsx():
    file = request.files.get("file")
    if not file:
        return error_response("file is required (.xlsx).", 422)
    if not str(file.filename or "").lower().endswith(".xlsx"):
        return error_response("Only .xlsx files are supported.", 422)

    replace_existing = (request.form.get("replace_existing") or "true").strip().lower() != "false"
    section_hint = (request.form.get("section") or "").strip().upper() or None
    if section_hint and not validate_course_section(section_hint):
        return error_response("section must follow the format AAAA X-X (e.g. BSIT 1-1).", 422)

    wb, err = _load_xlsx_workbook(file)
    if err:
        return error_response(err, 500)

    uploader_id = get_jwt_identity()
    batch = ImportBatch(
        id=_gen_id("IMP"),
        kind="schedules",
        filename=file.filename,
        section=section_hint,
        created_by=uploader_id,
        created_at=datetime.utcnow(),
    )
    db.session.add(batch)

    created = 0
    skipped = 0
    row_errors: list[dict] = []
    detected_sections: set[str] = set()

    parsed_rows: list[dict] = []

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = None
        col = {}
        for idx, row in enumerate(rows, start=1):
            if idx == 1:
                header = list(row or [])
                for j, h in enumerate(header):
                    key = _norm_header(h)
                    if key:
                        col[key] = j
                continue

            values = list(row or [])
            if not any(v is not None and str(v).strip() for v in values):
                continue

            def get(*keys):
                for k in keys:
                    j = col.get(k)
                    if j is None or j >= len(values):
                        continue
                    return values[j]
                return None

            section = str(get("section", "course_section", "course") or "").strip().upper()
            subject = str(get("subject") or "").strip()
            subject_code = str(get("subject_code", "code") or "").strip()
            day = _normalize_day(get("day"))
            room_key = str(get("room", "room_key", "room_id") or "").strip()
            start_raw = _coerce_schedule_time(get("start_time", "start"))
            end_raw = _coerce_schedule_time(get("end_time", "end"))
            campus_id = str(get("campus", "campus_id") or "").strip() or None
            building_id = str(get("building", "bldg", "building_id") or "").strip() or None

            if section:
                detected_sections.add(section)

            parsed_rows.append(
                {
                    "sheet": ws.title,
                    "row": idx,
                    "section": section,
                    "subject": subject,
                    "subject_code": subject_code,
                    "day": day,
                    "room_key": room_key,
                    "start_raw": start_raw,
                    "end_raw": end_raw,
                    "campus_id": campus_id,
                    "building_id": building_id,
                }
            )

    # Determine target section.
    target_section = section_hint or (next(iter(detected_sections)) if detected_sections else None)
    if not target_section:
        return error_response("No section detected. Provide section in the file or in form field 'section'.", 422)
    if len({s for s in detected_sections if s}) > 1 and not section_hint:
        return error_response(
            "Schedule upload contains multiple sections. Upload one section per file, or provide a fixed 'section' form field.",
            422,
        )
    if target_section and not validate_course_section(target_section):
        return error_response("Detected section is invalid (expected AAAA X-X).", 422)

    batch.section = target_section

    if replace_existing:
        # Delete existing schedules for this section (including prior imports).
        existing_ids = [s.id for s in Schedule.query.filter(Schedule.section == target_section).all()]
        if existing_ids:
            ScheduleImportItem.query.filter(ScheduleImportItem.schedule_id.in_(existing_ids)).delete(synchronize_session=False)
            Schedule.query.filter(Schedule.id.in_(existing_ids)).delete(synchronize_session=False)
            db.session.flush()

    for r in parsed_rows:
        if not r.get("section"):
            r["section"] = target_section
        if str(r.get("section") or "").strip().upper() != target_section:
            skipped += 1
            row_errors.append({"sheet": r.get("sheet"), "row": r.get("row"), "error": "section mismatch."})
            continue

        if not r.get("day"):
            skipped += 1
            row_errors.append({"sheet": r.get("sheet"), "row": r.get("row"), "error": "day is required."})
            continue
        if not r.get("room_key"):
            skipped += 1
            row_errors.append({"sheet": r.get("sheet"), "row": r.get("row"), "error": "room is required."})
            continue
        if not r.get("start_raw") or not r.get("end_raw"):
            skipped += 1
            row_errors.append({"sheet": r.get("sheet"), "row": r.get("row"), "error": "start_time and end_time are required."})
            continue

        s = Schedule(
            id=_gen_id("SCH"),
            section=target_section,
            subject=r.get("subject") or None,
            subject_code=r.get("subject_code") or None,
            day=r.get("day"),
            campus_id=r.get("campus_id"),
            building_id=r.get("building_id"),
            room_key=r.get("room_key"),
            start_time_raw=r.get("start_raw"),
            end_time_raw=r.get("end_raw"),
        )

        if s.start_minutes is None or s.end_minutes is None or s.start_minutes >= s.end_minutes:
            skipped += 1
            row_errors.append({"sheet": r.get("sheet"), "row": r.get("row"), "error": "invalid time range."})
            continue

        db.session.add(s)
        db.session.flush()
        db.session.add(ScheduleImportItem(batch_id=batch.id, schedule_id=s.id, row_num=int(r.get("row") or 0) or None))
        created += 1

    db.session.commit()

    return success_response(
        {
            "batch": batch.to_dict(),
            "created": created,
            "skipped": skipped,
            "errors": row_errors[:200],
        },
        "Schedule import complete.",
        201,
    )
